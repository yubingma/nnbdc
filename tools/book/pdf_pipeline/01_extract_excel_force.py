import openpyxl
import re
import os

def clean_word(raw_word):
    if not raw_word: return None
    raw_word = str(raw_word).strip()
    if not raw_word: return None
    
    # 将内部换行转为空格
    raw_word = raw_word.replace('\n', ' ')
    
    # 如果内容超过 40 个字符，很可能是含义或句子，跳过
    if len(raw_word) > 40: return None
    
    # 增加对 ~ 和 .. (两个点以上) 的截断逻辑
    raw_word = re.split(r'[\(（\[【\u2E80-\u9FFF/~]|\.\.', str(raw_word))[0].strip()
    
    # 物理净化：循环清除所有可能的词尾标记，直到没有变化
    while True:
        old_word = raw_word
        # 1. 清除末尾空格+任意词性标记
        raw_word = re.sub(r'\s+(adj|adv|prep|conj|pron|num|art|int|vt|vi|n|v|sb|sth|phr)\.*$', '', raw_word, flags=re.IGNORECASE)
        # 2. 强力裁掉黏连的 n, v, adj (针对特定后缀结尾的)
        if len(raw_word) > 5:
            # 形容词结尾清理
            if raw_word.endswith('adj'): raw_word = raw_word[:-3]
            elif raw_word.endswith('adv'): raw_word = raw_word[:-3]
            # 动词/名词结尾清理
            elif (raw_word.endswith('ionn') or raw_word.endswith('ancen') or raw_word.endswith('encen') or
                raw_word.endswith('mentn') or raw_word.endswith('ablen') or raw_word.endswith('nessn')):
                raw_word = raw_word[:-1]
            elif (raw_word.endswith('ancev') or raw_word.endswith('encev') or raw_word.endswith('ionv') or 
                  raw_word.endswith('atev') or raw_word.endswith('itev')):
                raw_word = raw_word[:-1]
        
        # 3. 再清理掉边缘标点，包含多余的点号
        raw_word = raw_word.strip('.,/*+~ ')
        if raw_word.endswith('..'): raw_word = raw_word[:-2].strip()
        
        if raw_word == old_word: break
        
        if raw_word == old_word: break

    clean_w = raw_word
    
    # 如果清理后只剩下词性缩写，干掉
    if clean_w.lower() in ['vt', 'vi', 'adj', 'adv', 'n', 'v', 'prep', 'phr', 'sb', 'sth', 'adj/n', 'n/v']:
        return None
    
    # 安全性检查：必须包含至少一个字母，且不能全是数字
    if not any(c.isalpha() for c in clean_w):
        return None
    
    # 如果经过裁剪后还是太长或者包含太多空格，可能不是个单词单词而是句子，过滤掉
    if len(clean_w.split()) > 3: # 例如 "in a way" 是合法的词组，但太长就不像词了
        return None
        
    return clean_w

def extract_ielts_force(input_path, output_path):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    words = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows(values_only=True):
            if not row: continue
            for val in row:
                cleaned = clean_word(val)
                if cleaned and len(cleaned) > 1:
                    words.append(cleaned)
    
    # 去重
    unique_words = sorted(list(set(words)))
    
    with open(output_path, 'w', encoding='utf-8') as f:
        for w in unique_words:
            f.write(w + "\n")
            
    print(f"🔥 暴力模式提取完成！")
    print(f"全书 63 个 Sheet 总计发现候选单词/词组: {len(words)} 个。")
    print(f"全局去重后获得最终词汇表: {len(unique_words)} 个。")
    print(f"保存至: {output_path}")

if __name__ == '__main__':
    extract_ielts_force('/Volumes/ssd/ppdc/tools/book/雅思/彩色词汇中英版.xlsx', '/Volumes/ssd/ppdc/tools/book/pdf_pipeline/raw_output.txt')
