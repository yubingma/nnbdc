import openpyxl
import re
import os

def clean_word(raw_word):
    if not raw_word: return None
    raw_word = str(raw_word).replace('\n', ' ').strip()
    # 去除词性标记
    tags = [r'\badj\.', r'\badv\.', r'\bn\.', r'\bv\.', r'\bprep\.', r'\bconj\.', r'\bpron\.', r'\bnum\.', r'\bart\.', r'\bint\.', r'\bphr\.']
    for tag in tags:
        raw_word = re.sub(tag, '', raw_word)
    
    # 彻底去除星号和其他非字母修饰
    raw_word = raw_word.replace('*', '').strip()
    
    # 直接按第一个括号、中文或特殊符号切断
    clean_w = re.split(r'[\(（\[【\u4e00-\u9fa5]', raw_word)[0].strip()
    # 转换成小写进行初步过滤，但保留原形态
    return clean_w

def extract_from_excel(input_path, output_path):
    wb = openpyxl.load_workbook(input_path, data_only=True)
    words = []
    
    # 遍历所有 Sheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        word_indices = [1, 5, 9]
        for row in ws.iter_rows(values_only=True):
            for idx in word_indices:
                if row and idx < len(row):
                    val = row[idx]
                    cleaned = clean_word(val)
                    if cleaned and len(cleaned) > 1: # 忽略单字母或空
                        words.append(cleaned)
    
    # 全局去重，并按照字母顺序排列，以便肉眼审计
    unique_words = sorted(list(set(words)))
    
    with open(output_path, 'w', encoding='utf-8') as f:
        for w in unique_words:
            f.write(w + "\n")
            
    print(f"✅ 全量 Excel 提取完成！遍历了 {len(wb.sheetnames)} 个 Sheet，")
    print(f"共发现 {len(words)} 条记录，去重后剩余 {len(unique_words)} 个单词。")
    print(f"最终结果已保存至: {output_path}")

if __name__ == '__main__':
    extract_from_excel('/Volumes/ssd/ppdc/tools/book/雅思/彩色词汇中英版.xlsx', '/Volumes/ssd/ppdc/tools/book/pdf_pipeline/raw_output.txt')
