import pdfplumber
import openpyxl
import re
import os
import hashlib

def get_md5(text):
    return hashlib.md5(text.encode('utf-8')).hexdigest()

def extract_pdf_words(pdf_path, output_txt):
    POS_PATTERN = r'\b([A-Za-z\-\']{2,})\s+(?:[a-z]{1,4}\.)'
    words = []
    seen = set()
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    matches = re.findall(POS_PATTERN, text)
                    for m in matches:
                        m_lower = m.lower()
                        if m_lower not in seen:
                            seen.add(m_lower)
                            words.append(m)
        
        with open(output_txt, 'w', encoding='utf-8') as f:
            for w in words:
                f.write(w + "\n")
        print(f"✅ Extracted {len(words)} words from {os.path.basename(pdf_path)}")
    except Exception as e:
        print(f"❌ Error processing {pdf_path}: {e}")

def save_data(data, output_dir, filename):
    data.sort(key=lambda x: (x['topic'], x['md5']))
    output_txt = os.path.join(output_dir, filename)
    with open(output_txt, 'w', encoding='utf-8') as f:
        for item in data:
            f.write(item['word'] + "\n")
    print(f"✅ Extracted {len(data)} entries -> {filename}")

def extract_excel_all_sheets(excel_path, output_dir):
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        level_sheets = ['A1', 'A2', 'B1', 'B2', 'C1', 'C2']
        phrase_sheets = ['idiom', '短语', '动词短语']
        
        # Mapping for display names in filenames
        sheet_display_names = {
            'idiom': '习语',
            '短语': '短语',
            '动词短语': '动词短语'
        }
        
        for sheet_name in wb.sheetnames:
            if sheet_name == '使用说明':
                continue
                
            ws = wb[sheet_name]
            header_row = 1
            header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
            if 'base word' not in header:
                header = [str(c.value).strip().lower() if c.value else "" for c in ws[2]]
                header_row = 2
            
            try:
                word_idx = header.index('base word')
            except ValueError:
                word_idx = 0
            try:
                level_idx = header.index('level')
            except ValueError:
                level_idx = -1
            try:
                topic_idx = header.index('topic')
            except ValueError:
                topic_idx = -1
            
            if sheet_name in level_sheets:
                data = []
                seen = set()
                for row in ws.iter_rows(min_row=header_row+1, values_only=True):
                    if row and row[word_idx]:
                        word = str(row[word_idx]).strip()
                        word_lower = word.lower()
                        if word_lower not in seen:
                            topic = str(row[topic_idx]).strip() if (topic_idx != -1 and topic_idx < len(row) and row[topic_idx]) else ""
                            seen.add(word_lower)
                            data.append({'word': word, 'topic': topic, 'md5': get_md5(word_lower)})
                save_data(data, output_dir, f"CERF等级词汇{sheet_name}.txt")
                
            elif sheet_name in phrase_sheets:
                level_data = {}
                display_name = sheet_display_names.get(sheet_name, sheet_name)
                for row in ws.iter_rows(min_row=header_row+1, values_only=True):
                    if row and row[word_idx]:
                        word = str(row[word_idx]).strip()
                        level = str(row[level_idx]).strip().upper() if (level_idx != -1 and level_idx < len(row) and row[level_idx]) else "UNKNOWN"
                        topic = str(row[topic_idx]).strip() if (topic_idx != -1 and topic_idx < len(row) and row[topic_idx]) else ""
                        if level not in level_data:
                            level_data[level] = []
                        level_data[level].append({'word': word, 'topic': topic, 'md5': get_md5(word.lower())})
                
                for level, data in level_data.items():
                    save_data(data, output_dir, f"CERF{display_name}_{level}.txt")
            
            elif sheet_name == 'CEFR-去重':
                data = []
                seen = set()
                for row in ws.iter_rows(min_row=header_row+1, values_only=True):
                    if row and row[word_idx]:
                        word = str(row[word_idx]).strip()
                        word_lower = word.lower()
                        if word_lower not in seen:
                            topic = str(row[topic_idx]).strip() if (topic_idx != -1 and topic_idx < len(row) and row[topic_idx]) else ""
                            seen.add(word_lower)
                            data.append({'word': word, 'topic': topic, 'md5': get_md5(word_lower)})
                save_data(data, output_dir, "CERF_CEFR-去重.txt")

    except Exception as e:
        print(f"❌ Error processing {excel_path}: {e}")

if __name__ == "__main__":
    base_dir = "/Volumes/ssd/ppdc/tools/book/多邻国/"
    output_dir = "/Volumes/ssd/ppdc/tools/book/多邻国/extracted/"
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # Process PDFs
    pdf_books = ["多邻国核心词汇初级.pdf", "多邻国核心词汇中级.pdf", "多邻国核心词汇专业.pdf"]
    for book in pdf_books:
        extract_pdf_words(os.path.join(base_dir, book), os.path.join(output_dir, book.replace('.pdf', '.txt')))
        
    # Process Excel
    extract_excel_all_sheets(os.path.join(base_dir, "7.CEFR 7243词单词表-re.xlsx"), output_dir)
