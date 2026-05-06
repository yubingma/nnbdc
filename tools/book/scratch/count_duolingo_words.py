import pdfplumber
import openpyxl
import re
import os

def count_pdf_words(pdf_path):
    # More inclusive POS pattern
    POS_PATTERN = r'\b([A-Za-z\-\']{2,})\s+(?:[a-z]{1,4}\.)'
    seen = set()
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    matches = re.findall(POS_PATTERN, text)
                    for m in matches:
                        seen.add(m.lower())
    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")
        return 0
    return len(seen)

def count_excel_words(excel_path):
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if 'CEFR-去重' in wb.sheetnames:
            ws = wb['CEFR-去重']
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    count += 1
            return count
        else:
            seen = set()
            for name in ['A1', 'A2', 'B1', 'B2', 'C1', 'C2']:
                if name in wb.sheetnames:
                    ws = wb[name]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and row[0]:
                            seen.add(str(row[0]).strip().lower())
            return len(seen)
    except Exception as e:
        print(f"Error processing {excel_path}: {e}")
        return 0

if __name__ == "__main__":
    base_dir = "/Volumes/ssd/ppdc/tools/book/多邻国/"
    books = [
        "多邻国核心词汇初级.pdf",
        "多邻国核心词汇中级.pdf",
        "多邻国核心词汇专业.pdf",
        "7.CEFR 7243词单词表-re.xlsx"
    ]
    
    total = 0
    results = {}
    for book in books:
        path = os.path.join(base_dir, book)
        if book.endswith(".pdf"):
            count = count_pdf_words(path)
        else:
            count = count_excel_words(path)
        results[book] = count
        total += count
    
    for book, count in results.items():
        print(f"{book}: {count}")
    
    print(f"\nTotal words: {total}")
